import requests
import json
import os
from typing import Dict, List
from collections import defaultdict
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class VikaCrawler:
    def __init__(self):
        self.api_token = "uskciRoNmrigKVACFTUZjJe"
        self.base_url = "https://api.vika.cn/fusion/v1"
        
    def get_datasheet(self, datasheet_id: str, view_id: str = None) -> Dict:
        """获取维格表数据"""
        url = f"{self.base_url}/datasheets/{datasheet_id}/records"
        headers = {
            "Authorization": f"Bearer {self.api_token}",
            "Content-Type": "application/json"
        }
        
        all_records = []
        page_size = 50
        page_num = 1
        
        while True:
            params = {
                "pageSize": page_size,
                "pageNum": page_num
            }
            if view_id:
                params["viewId"] = view_id
            
            try:
                if page_num > 1:
                    import time
                    time.sleep(1)
                
                response = requests.get(url, headers=headers, params=params)
                response.raise_for_status()
                
                data = response.json()
                
                if "data" in data and "records" in data["data"]:
                    records = data["data"]["records"]
                    all_records.extend(records)
                    print(f"成功获取第{page_num}页数据，共{len(records)}条记录")
                    
                    if len(records) < page_size:
                        break
                        
                    page_num += 1
                else:
                    print(f"响应数据格式异常: {data}")
                    break
                    
            except requests.exceptions.RequestException as e:
                print(f"请求失败: {str(e)}")
                if hasattr(e.response, 'text'):
                    print(f"错误响应: {e.response.text}")
                print(f"已获取 {len(all_records)} 条记录，停止获取更多数据")
                break
                
        return all_records

def calculate_workload(records: List[Dict]) -> Dict:
    """计算每个员工每月的工时统计"""
    workload_stats = defaultdict(lambda: defaultdict(float))
    
    print(f"\n总记录数: {len(records)}")
    
    for record in records:
        fields = record.get("fields", {})
        
        # 获取关键字段
        employee = fields.get("执行人名称", [])
        employee_names = []
        for emp in employee:
            if isinstance(emp, dict):
                name = emp.get("name", "")
                if name:
                    employee_names.append(name)
        
        # 使用实际完成时间而不是创建时间
        completion_timestamp = fields.get("实际完成时间")
        if completion_timestamp:
            date_str = datetime.fromtimestamp(completion_timestamp/1000).strftime("%Y-%m-%d")
        else:
            date_str = None
            
        status = fields.get("任务状态")
        actual_hours = fields.get("实际工时", 0)
        
        # 打印每条记录的关键信息
        print(f"\n记录信息:")
        print(f"任务名称: {fields.get('任务名称', '')}")
        print(f"执行人: {employee_names}")
        print(f"实际完成时间: {date_str}")
        print(f"状态: {status}")
        print(f"实际工时: {actual_hours}")
        
        # 检查状态值
        if status != "已完成":
            print(f"跳过：任务未完成 (状态值: {status})")
            continue
        if not employee_names:
            print("跳过：无执行人")
            continue
        if not date_str:
            print("跳过：无实际完成时间")
            continue
        if not actual_hours:
            print("跳过：无实际工时")
            continue
            
        try:
            completion_date = datetime.strptime(date_str, "%Y-%m-%d")
            month_key = completion_date.strftime("%Y-%m")
            
            hours_per_person = float(actual_hours) / len(employee_names)
            
            for emp_name in employee_names:
                workload_stats[emp_name][month_key] += hours_per_person
                print(f"添加工时: {emp_name} - {month_key} - {hours_per_person}小时")
                    
        except (ValueError, TypeError, AttributeError) as e:
            print(f"处理出错: {str(e)}")
            continue
            
    return workload_stats

def export_to_excel(workload_stats: Dict, total_hours: Dict, top_10_employees: List):
    """导出统计结果到Excel"""
    # 创建输出目录
    output_dir = "output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    wb = Workbook()
    
    # 创建总览sheet
    overview_sheet = wb.active
    overview_sheet.title = "工时统计总览"
    
    # 设置标题
    overview_sheet.cell(row=1, column=1, value='排名')
    overview_sheet.cell(row=1, column=2, value='执行人')
    overview_sheet.cell(row=1, column=3, value='总工时')
    
    # 设置标题样式
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for col in range(1, 4):
        cell = overview_sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 填充数据
    for i, (employee, hours) in enumerate(top_10_employees, 1):
        overview_sheet.cell(row=i+1, column=1, value=i)
        overview_sheet.cell(row=i+1, column=2, value=employee)
        overview_sheet.cell(row=i+1, column=3, value=round(hours, 1))
    
    # 调整列宽
    for col, width in [('A', 8), ('B', 20), ('C', 12)]:
        overview_sheet.column_dimensions[col].width = width
    
    # 创建月度明细sheet
    detail_sheet = wb.create_sheet("月度工时明细")
    
    # 获取所有月份
    all_months = sorted(set(month for stats in workload_stats.values() for month in stats.keys()))
    
    # 设置标题
    detail_sheet.cell(row=1, column=1, value='执行人')
    for col, month in enumerate(all_months, 2):
        detail_sheet.cell(row=1, column=col, value=month)
    detail_sheet.cell(row=1, column=len(all_months)+2, value='总计')
    
    # 设置标题样式
    for col in range(1, len(all_months)+3):
        cell = detail_sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 填充数据
    for row, (employee, total) in enumerate(top_10_employees, 2):
        detail_sheet.cell(row=row, column=1, value=employee)
        monthly_stats = workload_stats[employee]
        for col, month in enumerate(all_months, 2):
            hours = monthly_stats.get(month, 0)
            detail_sheet.cell(row=row, column=col, value=round(hours, 1))
        detail_sheet.cell(row=row, column=len(all_months)+2, value=round(total, 1))
    
    # 调整列宽
    detail_sheet.column_dimensions['A'].width = 20
    max_cols = min(len(all_months) + 2, 26)  # 限制最大列数为26（A-Z）
    for i in range(max_cols):
        col_letter = chr(ord('B') + i)
        if col_letter <= 'Z':  # 确保列名在A-Z范围内
            detail_sheet.column_dimensions[col_letter].width = 12
    
    # 保存文件
    filename = f"工时统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    # 打印完整的文件路径
    abs_path = os.path.abspath(filepath)
    print(f"\n文件已保存到: {abs_path}")
    
    return filepath

def main():
    try:
        crawler = VikaCrawler()
        datasheet_id = "dstBDj9lJLFBTpSQHB"
        
        print("\n=== 开始获取维格表数据 ===")
        records = crawler.get_datasheet(datasheet_id)
        
        if not records:
            print("警告：未获取到任何记录！")
            return
        
        print("\n=== 开始计算工时统计（按实际完成时间统计已完成任务的工时）===")
        workload_stats = calculate_workload(records)
        
        if not workload_stats:
            print("警告：没有找到符合条件的工时记录！")
            return
            
        # 计算总工时
        total_hours = defaultdict(float)
        for employee, monthly_stats in workload_stats.items():
            total_hours[employee] = sum(monthly_stats.values())
        
        # 获取前10名
        top_10_employees = sorted(total_hours.items(), key=lambda x: x[1], reverse=True)[:10]
        
        # 导出到Excel
        filename = export_to_excel(workload_stats, total_hours, top_10_employees)
        print(f"\n统计结果已导出到: {filename}")
        
        # 打印统计结果
        print("\n=== 总工时排名（前10名）===")
        for i, (employee, hours) in enumerate(top_10_employees, 1):
            print(f"{i}. {employee}: {hours:.1f}小时")
            
    except Exception as e:
        print(f"\n发生错误: {str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    main()